from django.shortcuts import redirect, render, get_object_or_404
from django.views.generic import (
    TemplateView,
    CreateView,
    ListView,
    DetailView,
    UpdateView,
    DeleteView,
    View,
)
from django.urls import reverse, reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.utils import timezone
from django.contrib import messages
from django.http import (
    HttpResponseRedirect,
    JsonResponse,
    HttpResponseBadRequest,
    HttpResponseForbidden,
    HttpResponse,
)
from django.core.serializers.json import DjangoJSONEncoder
import json
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import requests
import os
import datetime
import pytz
from django.db.models import Q, Count, Sum
from openpyxl import Workbook

from .models import Session, Service, Profile, Review, Payment
from .forms import ServiceForm, BookingForm, ReviewForm, SessionForm
from .utils.google_calendar import (
    create_coach_calendar_event,
    delete_coach_calendar_event,
)
from accounts.models import Profile


class HomeView(TemplateView):
    template_name = "home.html"


class SessionHistoryView(LoginRequiredMixin, ListView):
    model = Session
    template_name = "viewer/session_history.html"
    context_object_name = "object_list"
    paginate_by = 10

    def get_queryset(self):
        user_profile = self.request.user.profile
        if user_profile.is_coach:
            return Session.objects.filter(coach=user_profile).order_by("-date_time")
        else:
            return Session.objects.filter(client=user_profile).order_by("-date_time")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_profile = self.request.user.profile
        now = timezone.now()
        # Nadcházející sessions (všechny statusy)
        if user_profile.is_coach:
            context["upcoming_sessions"] = Session.objects.filter(
                coach=user_profile, date_time__gt=now
            ).order_by("date_time")
            context["past_sessions"] = Session.objects.filter(
                coach=user_profile, date_time__lte=now
            ).order_by("-date_time")
        else:
            context["upcoming_sessions"] = Session.objects.filter(
                client=user_profile, date_time__gt=now
            ).order_by("date_time")
            context["past_sessions"] = Session.objects.filter(
                client=user_profile, date_time__lte=now
            ).order_by("-date_time")
        # Přidám časové pásmo uživatele
        context["timezone"] = (
            user_profile.timezone if hasattr(user_profile, "timezone") else "UTC"
        )
        return context


class SessionDetailView(LoginRequiredMixin, DetailView):
    model = Session
    template_name = "viewer/session_detail.html"
    context_object_name = "session"

    def get_queryset(self):
        user_profile = self.request.user.profile
        if user_profile.is_coach:
            return Session.objects.filter(coach=user_profile)
        else:
            return Session.objects.filter(client=user_profile)

    def post(self, request, *args, **kwargs):
        session = self.get_object()
        user_profile = request.user.profile
        # Inline potvrzení pouze pro kouče a pouze pokud je session PENDING/CHANGED
        if user_profile.is_coach and session.status in ['PENDING', 'CHANGED']:
            meeting_url = request.POST.get('meeting_url', '').strip()
            meeting_address = request.POST.get('meeting_address', '').strip()
            errors = []
            if session.type == 'online':
                if not meeting_url:
                    errors.append('Online session must have a meeting link.')
                else:
                    session.meeting_url = meeting_url
            elif session.type == 'personal':
                if not meeting_address:
                    errors.append('Personal session must have an address.')
                else:
                    session.meeting_address = meeting_address
            if not errors:
                session.status = 'CONFIRMED'
                session.save()
                messages.success(request, 'Session has been confirmed.')
                return redirect('viewer:session_detail', pk=session.pk)
            else:
                for err in errors:
                    messages.error(request, err)
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = self.get_object()
        user_profile = self.request.user.profile
        can_edit_or_cancel = False
        if user_profile == session.client:
            seconds = (session.date_time - timezone.now()).total_seconds()
            if seconds > 24*60*60:
                can_edit_or_cancel = True
        context['can_edit_or_cancel'] = can_edit_or_cancel
        return context


class ServiceListView(ListView):
    model = Service
    template_name = "viewer/service_list.html"
    context_object_name = "services"

    def get_queryset(self):
        queryset = Service.objects.filter(is_active=True)
        if self.request.user.is_authenticated and hasattr(self.request.user, "profile"):
            if self.request.user.profile.is_coach:
                # Coaches see only their services
                return queryset.filter(coach=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_coach"] = (
            self.request.user.is_authenticated
            and hasattr(self.request.user, "profile")
            and self.request.user.profile.is_coach
        )
        return context


class ServiceDetailView(DetailView):
    model = Service
    template_name = "viewer/service_detail.html"
    context_object_name = "service"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            context["is_coach"] = (
                hasattr(self.request.user, "profile")
                and self.request.user.profile.is_coach
            )
            if not context["is_coach"]:
                # Get all confirmed sessions for this service
                booked_sessions = Session.objects.filter(
                    service=self.object, status="CONFIRMED"
                ).values_list("date_time", flat=True)

                # Get all confirmed sessions for the coach
                coach_sessions = (
                    Session.objects.filter(
                        coach=self.object.coach.profile, status="CONFIRMED"
                    )
                    .exclude(service=self.object)
                    .values_list("date_time", flat=True)
                )

                # Combine all booked times
                all_booked_times = list(booked_sessions) + list(coach_sessions)

                # Get today's date
                today = timezone.now().date()

                # Generate available slots for next 7 days
                available_slots = []
                for day in range(7):
                    current_date = today + timezone.timedelta(days=day)
                    for hour in range(9, 18):  # 9 AM to 6 PM
                        slot_time = timezone.make_aware(
                            timezone.datetime.combine(
                                current_date, timezone.time(hour=hour)
                            )
                        )

                        # Check if this slot is available
                        is_available = True
                        for booked_time in all_booked_times:
                            if (
                                abs((slot_time - booked_time).total_seconds()) < 3600
                            ):  # 1 hour
                                is_available = False
                                break

                        if is_available:
                            available_slots.append(slot_time)

                context["available_slots"] = available_slots
                context["booked_slots"] = all_booked_times
        return context


class ServiceCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Service
    form_class = ServiceForm
    template_name = "viewer/service_form.html"
    success_url = reverse_lazy("viewer:services")

    def test_func(self):
        return (
            hasattr(self.request.user, "profile") and self.request.user.profile.is_coach
        )

    def form_valid(self, form):
        form.instance.coach = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Create New Service"
        context["button_text"] = "Create Service"
        return context


class ServiceUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Service
    form_class = ServiceForm
    template_name = "viewer/service_form.html"
    success_url = reverse_lazy("viewer:services")

    def test_func(self):
        service = self.get_object()
        return self.request.user == service.coach

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Edit Service"
        context["button_text"] = "Update Service"
        return context


class ServiceDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Service
    template_name = "viewer/service_confirm_delete.html"
    success_url = reverse_lazy("viewer:services")

    def test_func(self):
        service = self.get_object()
        return self.request.user == service.coach


class BookingCreateView(LoginRequiredMixin, CreateView):
    model = Session
    form_class = BookingForm
    template_name = "viewer/booking_form.html"
    success_url = reverse_lazy("viewer:session_history")

    def dispatch(self, request, *args, **kwargs):
        if not request.user.profile.is_client:
            return HttpResponseForbidden("Only clients can create bookings.")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        # Pokud je v GET parametr service, předej ho do initial
        service = self.request.GET.get("service")
        if service:
            kwargs["initial"] = kwargs.get("initial", {})
            kwargs["initial"]["service"] = service
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Přidáme services_json pro JavaScript
        services = {}
        for service in Service.objects.filter(is_active=True):
            services[str(service.pk)] = {
                "price": str(service.price),
                "currency": service.currency,
                "duration": service.duration,
                "description": service.description,
                "session_type": service.session_type,
            }
        context["services_json"] = json.dumps(services)
        return context

    def form_valid(self, form):
        form.instance.client = self.request.user.profile
        service = form.instance.service
        coach_user = service.coach
        form.instance.status = "PENDING"
        form.instance.duration = service.duration
        form.instance.type = service.session_type  # Nastaví typ podle servisu
        try:
            coach_profile = Profile.objects.get(user=coach_user)
            form.instance.coach = coach_profile
            # Vytvoření události v Google kalendáři
            event = create_coach_calendar_event(
                coach_profile=coach_profile,
                summary=f"Session with {self.request.user.get_full_name()}",
                description=f"Service: {service.name}\nClient: {self.request.user.get_full_name()}",
                start_dt=form.instance.date_time,
                end_dt=form.instance.date_time
                + timezone.timedelta(minutes=service.duration),
                timezone_str=str(timezone.get_current_timezone()),
            )
            if event:
                form.instance.google_calendar_event_id = event["id"]
                messages.success(
                    self.request,
                    f'Session booked successfully! Google Calendar event created: {event.get("htmlLink")}',
                )
            else:
                messages.success(self.request, "Session booked successfully!")
        except Profile.DoesNotExist:
            raise Exception(f"Coach {coach_user} nemá profil!")
        except Exception as e:
            print(f"Google Calendar error: {e}")
            messages.warning(
                self.request,
                f"Session booked, but failed to create Google Calendar event: {str(e)}",
            )
        response = super().form_valid(form)
        
        # Vytvoření platby
        Payment.objects.create(
            session=self.object,
            amount=service.price,
            payment_method=form.cleaned_data["payment_method"],
        )
        return response


class SessionUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Session
    template_name = "viewer/booking_form.html"
    success_url = reverse_lazy("viewer:session_history")

    def get_form_class(self):
        user = self.request.user
        if user.profile.is_client:
            return BookingForm
        return SessionForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        # Předvyplníme coach z původní session nebo ze service
        if self.object and not kwargs.get('data'):
            initial = kwargs.get('initial', {})
            initial['coach'] = self.object.coach.id if self.object.coach else self.object.service.coach.profile.id
            kwargs['initial'] = initial
        return kwargs

    def test_func(self):
        session = self.get_object()
        user = self.request.user
        if user.is_superuser:
            return True
        if user.profile.is_client and session.client.user == user:
            return True
        if user.profile.is_coach and session.coach.user == user:
            return True
        return False

    def form_valid(self, form):
        import datetime

        session = form.instance
        dt = form.cleaned_data.get("date_time")
        if isinstance(dt, str):
            try:
                dt = datetime.datetime.strptime(dt, "%Y-%m-%d %H:%M")
                dt = timezone.make_aware(dt)
            except Exception:
                pass
        if dt:
            session.date_time = dt

        # Kontrola meeting_url a meeting_address pouze pro CONFIRMED session
        if session.status == "CONFIRMED":
            if form.cleaned_data.get('type') == "online" and not form.cleaned_data.get('meeting_url'):
                form.add_error("meeting_url", "Online session must have a meeting link.")
                return self.form_invalid(form)
            if form.cleaned_data.get('type') == "personal" and not form.cleaned_data.get('meeting_address'):
                form.add_error("meeting_address", "Personal session must have an address.")
                return self.form_invalid(form)

        if self.request.user.profile.is_coach and self.request.POST.get(
            "confirm_and_save"
        ):
            session.status = "CONFIRMED"
        elif self.request.user.profile.is_coach and session.status == "PENDING":
            payment = session.payments.first()
            if payment and payment.paid_at:
                session.status = "CONFIRMED"
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["services"] = Service.objects.filter(is_active=True)
        context["title"] = "Edit Session"
        context["button_text"] = "Update Session"
        context["is_edit"] = True
        context["current_time"] = timezone.now()
        if self.object and self.object.service:
            context["price"] = self.object.service.price
            context["duration"] = self.object.service.duration
            context["description"] = self.object.service.description
        # Přidám services_json pro JS sloty
        services = {}
        for service in Service.objects.filter(is_active=True):
            services[str(service.pk)] = {
                "price": str(service.price),
                "currency": service.currency,
                "duration": service.duration,
                "description": service.description,
                "session_type": service.session_type,
            }
        context["services_json"] = json.dumps(services)
        # Přidej initial_date_time pro JS - upraveno na ISO format
        if self.object and self.object.date_time:
            context["initial_date_time"] = self.object.date_time.astimezone(
                timezone.get_current_timezone()
            ).isoformat()
        else:
            context["initial_date_time"] = ""
        # Přidám rozdíl v hodinách do kontextu
        user = self.request.user
        user_tz = timezone.get_current_timezone()
        if hasattr(user, "profile") and getattr(user.profile, "timezone", None):
            try:
                user_tz = pytz.timezone(user.profile.timezone)
            except Exception:
                pass
        if self.object and self.object.date_time:
            now = timezone.now().astimezone(user_tz)
            session_time = self.object.date_time.astimezone(user_tz)
            diff = (session_time - now).total_seconds() / 3600
            context["hours_until_session"] = diff
        else:
            context["hours_until_session"] = None
        return context


class SessionCancelView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        session = Session.objects.get(pk=self.kwargs["pk"])
        user = self.request.user
        return (
            user.profile == session.client
            or user.profile == session.coach
            or user.is_superuser
        )

    def post(self, request, *args, **kwargs):
        session = Session.objects.get(pk=self.kwargs["pk"])
        if session.can_cancel:
            # Pokud existuje ID události v kalendáři, smaž ji
            if session.google_calendar_event_id:
                try:
                    delete_coach_calendar_event(
                        session.coach, session.google_calendar_event_id
                    )
                except Exception as e:
                    print(f"Failed to delete Google Calendar event: {e}")
                    messages.warning(
                        request,
                        "Session cancelled, but failed to delete Google Calendar event.",
                    )

            session.status = "CANCELLED"
            session.save()
            messages.success(request, "Session has been cancelled successfully.")
        else:
            messages.error(
                request, "Cannot cancel session less than 24 hours before start time."
            )
        return redirect("viewer:session_detail", pk=session.pk)


class ReviewCreateView(LoginRequiredMixin, CreateView):
    model = Review
    form_class = ReviewForm
    template_name = "viewer/review_form.html"

    def form_valid(self, form):
        session_id = self.kwargs["pk"]
        form.instance.session_id = session_id
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("viewer:session_detail", args=[self.kwargs["pk"]])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["session"] = Session.objects.get(pk=self.kwargs["pk"])
        return context


def home(request):
    return render(request, "home.html")


class CreatePayPalOrderView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        session_id = request.POST.get("session_id")
        if not session_id:
            return HttpResponseBadRequest("Missing session_id")
        try:
            session = Session.objects.get(pk=session_id)
            payment = session.payments.first()
            if not payment:
                return HttpResponseBadRequest("No payment found for this session")
        except Session.DoesNotExist:
            return HttpResponseBadRequest("Session not found")

        # PayPal credentials from .env
        PAYPAL_CLIENT_ID = os.environ.get("PAYPAL_CLIENT_ID")
        PAYPAL_CLIENT_SECRET = os.environ.get("PAYPAL_CLIENT_SECRET")
        PAYPAL_API_BASE = (
            "https://api-m.sandbox.paypal.com"  # change to live for production
        )

        # Get access token
        auth_response = requests.post(
            f"{PAYPAL_API_BASE}/v1/oauth2/token",
            auth=(PAYPAL_CLIENT_ID, PAYPAL_CLIENT_SECRET),
            data={"grant_type": "client_credentials"},
        )
        if auth_response.status_code != 200:
            return JsonResponse({"error": "PayPal auth failed"}, status=500)
        access_token = auth_response.json()["access_token"]

        # Create order
        order_data = {
            "intent": "CAPTURE",
            "purchase_units": [
                {
                    "amount": {
                        "currency_code": (
                            payment.payment_method.name.upper()
                            if payment.payment_method.name.upper()
                            in ["USD", "EUR", "CZK"]
                            else "USD"
                        ),
                        "value": str(payment.amount),
                    },
                    "description": f"Session: {session.service.name} ({session.date_time})",
                }
            ],
            "application_context": {
                "brand_name": "LifeCoach",
                "locale": "en-US",
                "return_url": request.build_absolute_uri(
                    f"/paypal/return/{session.id}/"
                ),
                "cancel_url": request.build_absolute_uri(
                    f"/paypal/cancel/{session.id}/"
                ),
                "user_action": "PAY_NOW",
            },
        }
        order_response = requests.post(
            f"{PAYPAL_API_BASE}/v2/checkout/orders",
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {access_token}",
            },
            json=order_data,
        )
        if order_response.status_code != 201:
            return JsonResponse(
                {
                    "error": "PayPal order creation failed",
                    "details": order_response.json(),
                },
                status=500,
            )
        order = order_response.json()
        approval_url = next(
            (l["href"] for l in order["links"] if l["rel"] == "approve"), None
        )
        if not approval_url:
            return JsonResponse({"error": "No approval url from PayPal"}, status=500)
        # Optionally save order_id to payment
        payment.transaction_id = order["id"]
        payment.save()
        return JsonResponse({"approval_url": approval_url})


class PayPalReturnView(LoginRequiredMixin, View):
    def get(self, request, session_id):
        from django.contrib import messages

        session = Session.objects.get(pk=session_id)
        payment = session.payments.first()
        order_id = payment.transaction_id
        # PayPal credentials
        PAYPAL_CLIENT_ID = os.environ.get("PAYPAL_CLIENT_ID")
        PAYPAL_CLIENT_SECRET = os.environ.get("PAYPAL_CLIENT_SECRET")
        PAYPAL_API_BASE = "https://api-m.sandbox.paypal.com"
        # Get access token
        auth_response = requests.post(
            f"{PAYPAL_API_BASE}/v1/oauth2/token",
            auth=(PAYPAL_CLIENT_ID, PAYPAL_CLIENT_SECRET),
            data={"grant_type": "client_credentials"},
        )
        access_token = auth_response.json()["access_token"]
        # Get order details
        order_response = requests.get(
            f"{PAYPAL_API_BASE}/v2/checkout/orders/{order_id}",
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {access_token}",
            },
        )
        order = order_response.json()
        if order.get("status") == "COMPLETED" or order.get("status") == "APPROVED":
            payment.paid_at = timezone.now()
            payment.save()
            messages.success(request, "Payment successful!")
        else:
            messages.warning(request, "Payment not completed. Please try again.")
        return redirect("viewer:session_detail", pk=session_id)


class PayPalCancelView(LoginRequiredMixin, View):
    def get(self, request, session_id):
        from django.contrib import messages

        messages.warning(request, "Payment was cancelled.")
        return redirect("viewer:session_detail", pk=session_id)


class MarkAsPaidView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        session = get_object_or_404(Session, pk=self.kwargs["pk"])
        return self.request.user.profile == session.coach

    def post(self, request, pk):
        session = get_object_or_404(Session, pk=pk)
        payment = session.payments.first()
        if payment and not payment.paid_at:
            payment.paid_at = timezone.now()
            payment.save()
            messages.success(request, "Session marked as paid successfully.")
        else:
            messages.error(request, "Could not mark session as paid.")
        return redirect("viewer:session_detail", pk=pk)


class AvailableSlotsView(LoginRequiredMixin, View):
    def get(self, request):
        service_id = request.GET.get("service")
        if not service_id:
            return JsonResponse({"error": "Service ID is required"}, status=400)

        try:
            service = Service.objects.get(pk=service_id)

            # Get user's timezone
            user_timezone = timezone.get_current_timezone()
            if request.user.profile.timezone:
                try:
                    user_timezone = pytz.timezone(request.user.profile.timezone)
                except pytz.exceptions.UnknownTimeZoneError:
                    pass

            # Get current time in user's timezone
            now = timezone.now().astimezone(user_timezone)
            today = now.date()

            # Get all booked sessions for this coach
            booked_sessions = Session.objects.filter(
                Q(coach=service.coach.profile) | Q(client=request.user.profile),
                status__in=["CONFIRMED", "PENDING"],
                date_time__gte=now,
            )

            # Create a list of all booked times
            all_booked_times = []
            for session in booked_sessions:
                session_start = session.date_time.astimezone(user_timezone)
                session_end = session_start + timezone.timedelta(
                    minutes=session.duration
                )
                all_booked_times.append((session_start, session_end))

            slots = []
            # Generate available slots for the next 7 days
            for day in range(7):
                current_date = today + timezone.timedelta(days=day)

                # Skip if it's today and already past working hours
                if day == 0 and now.hour >= 17:
                    continue

                # Generate slots for each hour from 9 AM to 5 PM
                for hour in range(9, 17):  # 9:00 - 16:00 (last session starts at 16:00)
                    # Skip past hours for today
                    if day == 0 and hour <= now.hour:
                        continue

                    slot_time = timezone.make_aware(
                        datetime.datetime.combine(
                            current_date, datetime.time(hour=hour)
                        ),
                        timezone=user_timezone,
                    )

                    # Check if this slot is available
                    is_available = True
                    slot_end = slot_time + timezone.timedelta(minutes=service.duration)

                    # Check for overlaps with booked sessions
                    for booked_start, booked_end in all_booked_times:
                        if (slot_time < booked_end) and (slot_end > booked_start):
                            is_available = False
                            break

                    if is_available:
                        # Format the slot time in user's timezone
                        slot_local = slot_time.astimezone(user_timezone)
                        slots.append(
                            {
                                "value": slot_local.isoformat(),
                                "display": slot_local.strftime("%A %d.%m.%Y %H:%M"),
                            }
                        )

            return JsonResponse({"slots": slots})

        except Service.DoesNotExist:
            return JsonResponse({"error": "Service not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


class CoachReportView(LoginRequiredMixin, TemplateView):
    template_name = "viewer/coach_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        coach_profile = user.profile

        # Počet rezervací podle služby
        service_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("service__name")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Počet rezervací podle kategorie
        category_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("service__category__name")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Přehled plateb
        payments = Payment.objects.filter(session__coach=coach_profile)
        total_paid = (
            payments.filter(paid_at__isnull=False).aggregate(Sum("amount"))[
                "amount__sum"
            ]
            or 0
        )
        total_unpaid = (
            payments.filter(paid_at__isnull=True).aggregate(Sum("amount"))[
                "amount__sum"
            ]
            or 0
        )

        payment_methods = (
            payments.values("payment_method__name")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-count")
        )

        # Nejaktivnější klienti
        client_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("client__user__username")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Průměrné hodnocení služeb
        from django.db.models import Avg

        service_ratings = (
            Review.objects.filter(session__coach=coach_profile)
            .values("session__service__name")
            .annotate(avg_rating=Avg("rating"), count=Count("id"))
            .order_by("-avg_rating")
        )

        # Rezervace podle stavu
        status_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("status")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Přidám mapu názvu služby na ID pro odkazy na review
        service_id_map = {
            s.name: s.id for s in Service.objects.filter(coach=coach_profile.user)
        }
        context["service_id_map"] = service_id_map

        # Zrušené a zaplacené session
        cancelled_paid_sessions = Session.objects.filter(
            coach=coach_profile,
            status="CANCELLED",
            payments__paid_at__isnull=False
        ).select_related("client", "service").distinct()
        context["cancelled_paid_sessions"] = cancelled_paid_sessions

        context.update(
            {
                "service_counts": service_counts,
                "category_counts": category_counts,
                "total_paid": total_paid,
                "total_unpaid": total_unpaid,
                "payment_methods": payment_methods,
                "client_counts": client_counts,
                "service_ratings": service_ratings,
                "status_counts": status_counts,
            }
        )
        return context


class CoachReportExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        coach_profile = user.profile

        # Data stejně jako v CoachReportView
        service_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("service__name")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        category_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("service__category__name")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        status_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("status")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        client_counts = (
            Session.objects.filter(coach=coach_profile)
            .values("client__user__username")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        from django.db.models import Avg

        service_ratings = (
            Review.objects.filter(session__coach=coach_profile)
            .values("session__service__name")
            .annotate(avg_rating=Avg("rating"), count=Count("id"))
            .order_by("-avg_rating")
        )
        payments = Payment.objects.filter(session__coach=coach_profile)
        total_paid = (
            payments.filter(paid_at__isnull=False).aggregate(Sum("amount"))[
                "amount__sum"
            ]
            or 0
        )
        total_unpaid = (
            payments.filter(paid_at__isnull=True).aggregate(Sum("amount"))[
                "amount__sum"
            ]
            or 0
        )
        payment_methods = (
            payments.values("payment_method__name")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-count")
        )

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Service Usage"
        ws1.append(["Service", "Reservations"])
        for row in service_counts:
            ws1.append([row["service__name"], row["count"]])

        ws2 = wb.create_sheet(title="Categories")
        ws2.append(["Category", "Reservations"])
        for row in category_counts:
            ws2.append([row["service__category__name"] or "-", row["count"]])

        ws3 = wb.create_sheet(title="Status")
        ws3.append(["Status", "Count"])
        for row in status_counts:
            ws3.append([row["status"], row["count"]])

        ws4 = wb.create_sheet(title="Clients")
        ws4.append(["Client", "Reservations"])
        for row in client_counts:
            ws4.append([row["client__user__username"], row["count"]])

        ws5 = wb.create_sheet(title="Service Ratings")
        ws5.append(["Service", "Average Rating", "Reviews"])
        for row in service_ratings:
            ws5.append([row["session__service__name"], row["avg_rating"], row["count"]])

        ws6 = wb.create_sheet(title="Payments")
        ws6.append(["Total paid", total_paid])
        ws6.append(["Total unpaid", total_unpaid])
        ws6.append([])
        ws6.append(["Method", "Count", "Total"])
        for row in payment_methods:
            ws6.append([row["payment_method__name"], row["count"], row["total"]])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=coach_report.xlsx"
        wb.save(response)
        return response


class ServiceReviewListView(LoginRequiredMixin, ListView):
    template_name = "viewer/service_review_list.html"
    context_object_name = "reviews"

    def get_queryset(self):
        service_id = self.kwargs["service_id"]
        return Review.objects.filter(session__service_id=service_id).select_related(
            "session", "session__client"
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from viewer.models import Service

        context["service"] = Service.objects.get(pk=self.kwargs["service_id"])
        return context
